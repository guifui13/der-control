from io import BytesIO
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MACRO_ORDEM = [
    "Geometria", "Geotecnia", "Notas de serviço", "Terraplenagem",
    "Drenagem", "Pavimentação", "Sinalização", "Desvio de Tráfego",
    "Estruturas", "Meio ambiente", "Interferências", "Desapropriação",
    "Geral", "Estudos de tráfego", "Topografia",
]

MAPA_MACRO = {
    "A": "Geral", "B": "Desvio de Tráfego", "C": "Estruturas",
    "D": "Desapropriação", "F": "Geometria", "G": "Geotecnia",
    "H": "Drenagem", "I": "Interferências", "J": "Estudos de tráfego",
    "L": "Sinalização", "P": "Pavimentação", "Q": "Terraplenagem",
    "S": "Meio ambiente", "T": "Topografia",
}

CODIGOS_NOTA_SERVICO = ["P01", "Q01"]


def normalizar_codigo(codigo):
    if pd.isna(codigo):
        return ""
    codigo = str(codigo).strip()
    if codigo == "" or codigo.lower() == "nan":
        return ""
    if "/" not in codigo:
        partes = codigo.rsplit("-", 1)
        if len(partes) == 2:
            codigo = partes[0] + "/" + partes[1]
    return codigo


def limpar_nome_aba(nome):
    nome = str(nome)
    for char in ["\\", "/", "*", "[", "]", ":", "?"]:
        nome = nome.replace(char, "-")
    return nome[:31]


def extrair_disciplina(codigo):
    match = re.search(r"-([A-Z]\d{2})[/-]", str(codigo))
    return match.group(1) if match else ""


def obter_macrodisciplina(codigo):
    disciplina = extrair_disciplina(codigo)
    if disciplina in CODIGOS_NOTA_SERVICO:
        return "Notas de serviço"
    if disciplina:
        return MAPA_MACRO.get(disciplina[0], "Geral")
    return "Geral"


def texto_limpo(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def classificar_documento(status_k, status_o):
    k = texto_limpo(status_k).upper()
    o = texto_limpo(status_o).upper()

    concepcao = ressalva = aprovado = finalizado = 0
    status_resumo = "Inserir Documento"

    if any(t in k or t in o for t in ["FINALIZADO", "ASSINATURA DIGITAL", "ENTREGA FINAL"]):
        concepcao, ressalva, aprovado, finalizado = 0.30, 0.40, 0.20, 0.10
        status_resumo = "Finalizado"
    elif "RESSALVA" in o:
        concepcao, ressalva = 0.30, 0.40
        status_resumo = "Aprovado com ressalva"
    elif (("APROVADO" in o and "NÃO APROVADO" not in o and "NAO APROVADO" not in o)
          or any(t in k or t in o for t in ["AGUARDANDO ASSINATURA", "EMITIR APROVADO"])):
        concepcao, ressalva, aprovado = 0.30, 0.40, 0.20
        status_resumo = "Aprovado"
    elif k != "" and "INSERIR" not in k:
        concepcao = 0.30
        status_resumo = "Entregue"

    total = concepcao + ressalva + aprovado + finalizado
    return {
        "concepcao": concepcao,
        "ressalva": ressalva,
        "aprovado": aprovado,
        "finalizado": finalizado,
        "total": total,
        "status_resumo": status_resumo,
    }


def gerar_medicao(arquivo_eclic, arquivo_controle, abas_escolhidas=None):
    df_eclic = pd.read_excel(arquivo_eclic)
    base_eclic = {}

    for _, row in df_eclic.iterrows():
        codigo = normalizar_codigo(row.iloc[0])
        if codigo == "":
            continue
        base_eclic[codigo] = {
            "revisao": row.iloc[4] if len(row) > 4 else "",
            "status_k": row.iloc[10] if len(row) > 10 else "",
            "data_inicio_fluxo": row.iloc[11] if len(row) > 11 else "",
            "data_fim_fluxo": row.iloc[12] if len(row) > 12 else "",
            "status_o": row.iloc[14] if len(row) > 14 else "",
            "data_importacao": row.iloc[19] if len(row) > 19 else "",
        }

    xls_controle = pd.ExcelFile(arquivo_controle)
    if not abas_escolhidas:
        abas_escolhidas = xls_controle.sheet_names

    fill_titulo_disciplina = PatternFill("solid", fgColor="D9EAD3")
    fill_cabecalho = PatternFill("solid", fgColor="FFF2CC")
    fill_cabecalho_azul = PatternFill("solid", fgColor="BDD7EE")
    fill_total = PatternFill("solid", fgColor="D9EAD3")
    fill_resumo = PatternFill("solid", fgColor="FFF2CC")
    fill_resumo_total = PatternFill("solid", fgColor="FCE4D6")
    font_titulo = Font(bold=True, size=13)
    font_bold = Font(bold=True)
    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    wb = Workbook()
    wb.remove(wb.active)
    nao_encontrados = []
    resumo_por_aba = []

    for nome_aba in abas_escolhidas:
        df_controle = pd.read_excel(arquivo_controle, sheet_name=nome_aba)
        nome_aba_excel = limpar_nome_aba(nome_aba)
        ws = wb.create_sheet(nome_aba_excel)
        linha = 1
        grupos = {macro: [] for macro in MACRO_ORDEM}

        for _, row in df_controle.iterrows():
            if len(row) < 5:
                continue
            codigo = normalizar_codigo(row.iloc[4])
            if codigo == "":
                continue
            if codigo not in base_eclic:
                nao_encontrados.append({"aba": nome_aba, "codigo": codigo})
                continue

            info = base_eclic[codigo]
            classificacao = classificar_documento(info["status_k"], info["status_o"])
            macro = obter_macrodisciplina(codigo)
            grupos[macro].append([
                codigo, info["revisao"], info["data_inicio_fluxo"], info["data_fim_fluxo"],
                info["data_importacao"], info["status_k"], classificacao["concepcao"],
                classificacao["ressalva"], classificacao["aprovado"], classificacao["finalizado"],
                "", "", "", "", classificacao["total"], 1, classificacao["status_resumo"],
            ])

        cont_status = {"Inserir Documento": 0, "Entregue": 0, "Aprovado com ressalva": 0, "Aprovado": 0, "Finalizado": 0}
        total_contribuicao = 0
        total_docs = 0

        for macro in MACRO_ORDEM:
            documentos = grupos.get(macro, [])
            if not documentos:
                continue

            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=17)
            cell = ws.cell(linha, 1)
            cell.value = macro.upper()
            cell.font = font_titulo
            cell.fill = fill_titulo_disciplina
            cell.alignment = Alignment(horizontal="center", vertical="center")
            linha += 1

            cabecalhos = ["CÓDIGO", "REV", "DATA INÍCIO FLUXO", "DATA FIM FLUXO", "DATA IMPORTAÇÃO", "STATUS", "CONCEPÇÃO", "APROVADO C/ RESSALVA", "APROVADO", "FINALIZADO", "", "", "", "", "TOTAL", "QTD", "STATUS RESUMO"]
            for col, titulo in enumerate(cabecalhos, start=1):
                c = ws.cell(linha, col)
                c.value = titulo
                c.font = font_bold
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = borda
                c.fill = fill_cabecalho_azul if col in [7, 8, 9, 10] else fill_cabecalho
            linha += 1
            inicio_docs = linha

            for doc in documentos:
                cont_status[doc[16]] = cont_status.get(doc[16], 0) + 1
                total_contribuicao += float(doc[14] or 0)
                total_docs += 1
                for col, valor in enumerate(doc, start=1):
                    c = ws.cell(linha, col)
                    c.value = valor
                    c.border = borda
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    if col in [7, 8, 9, 10, 15]:
                        c.number_format = "0%"
                    if col in [3, 4, 5]:
                        c.number_format = "dd/mm/yyyy hh:mm:ss"
                linha += 1

            fim_docs = linha - 1
            ws.cell(linha, 1).value = "TOTAL DISCIPLINA"
            ws.cell(linha, 1).font = font_bold
            ws.cell(linha, 15).value = f"=IF(SUM(P{inicio_docs}:P{fim_docs})=0,0,SUM(O{inicio_docs}:O{fim_docs})/SUM(P{inicio_docs}:P{fim_docs}))"
            ws.cell(linha, 15).number_format = "0.00%"
            ws.cell(linha, 16).value = f"=SUM(P{inicio_docs}:P{fim_docs})"
            for col in range(1, 18):
                c = ws.cell(linha, col)
                c.fill = fill_total
                c.border = borda
                c.font = font_bold
            linha += 3

        larguras = {"A": 45, "B": 10, "C": 20, "D": 20, "E": 20, "F": 35, "G": 14, "H": 22, "I": 14, "J": 14, "K": 5, "L": 5, "M": 5, "N": 5, "O": 12, "P": 8, "Q": 22}
        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura
        ws.freeze_panes = "A1"

        resumo_por_aba.append({
            "aba": nome_aba_excel,
            "inserir": cont_status.get("Inserir Documento", 0),
            "entregue": cont_status.get("Entregue", 0),
            "ressalva": cont_status.get("Aprovado com ressalva", 0),
            "aprovado": cont_status.get("Aprovado", 0),
            "finalizado": cont_status.get("Finalizado", 0),
            "total": total_docs,
            "contribuicao": total_contribuicao,
            "percentual_medido": (total_contribuicao / total_docs) if total_docs else 0,
        })

    ws = wb.create_sheet("RESUMO_GERAL", 0)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    c = ws.cell(1, 1)
    c.value = "RESUMO GERAL - MEDIÇÃO E-CLIC PRODUTO REALIZADO"
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill_titulo_disciplina

    ws.cell(2, 3).value = 0.30
    ws.cell(2, 4).value = 0.70
    ws.cell(2, 5).value = 0.90
    ws.cell(2, 6).value = 1.00
    for col in range(3, 7):
        ws.cell(2, col).number_format = "0%"
        ws.cell(2, col).font = font_bold
        ws.cell(2, col).alignment = Alignment(horizontal="center")
        ws.cell(2, col).border = borda

    cab = ["ABA", "INSERIR DOCUMENTO", "ENTREGUE", "APROVADO C/ RESSALVA", "APROVADO", "FINALIZADO", "QTDE TOTAL", "CONTRIBUIÇÃO TOTAL", "% MEDIDO", "% INSERIR", "% ENTREGUE", "% APROVADO C/ RESSALVA", "% FINALIZADO"]
    for col, titulo in enumerate(cab, start=1):
        c = ws.cell(3, col)
        c.value = titulo
        c.font = font_bold
        c.fill = fill_resumo
        c.border = borda
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    linha = 4
    for item in resumo_por_aba:
        ws.cell(linha, 1).value = item["aba"]
        ws.cell(linha, 2).value = item["inserir"]
        ws.cell(linha, 3).value = item["entregue"]
        ws.cell(linha, 4).value = item["ressalva"]
        ws.cell(linha, 5).value = item["aprovado"]
        ws.cell(linha, 6).value = item["finalizado"]
        ws.cell(linha, 7).value = item["total"]
        ws.cell(linha, 8).value = item["contribuicao"]
        ws.cell(linha, 9).value = item["percentual_medido"]
        ws.cell(linha, 10).value = item["inserir"] / item["total"] if item["total"] else 0
        ws.cell(linha, 11).value = item["entregue"] / item["total"] if item["total"] else 0
        ws.cell(linha, 12).value = item["ressalva"] / item["total"] if item["total"] else 0
        ws.cell(linha, 13).value = item["finalizado"] / item["total"] if item["total"] else 0
        for col in range(1, 14):
            c = ws.cell(linha, col)
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col in [9, 10, 11, 12, 13]:
                c.number_format = "0.00%"
            if col == 8:
                c.number_format = "0.00"
        linha += 1

    ws.cell(linha, 1).value = "TOTAL GERAL"
    for col in range(2, 8):
        ws.cell(linha, col).value = f"=SUM({ws.cell(4, col).coordinate}:{ws.cell(linha-1, col).coordinate})"
    ws.cell(linha, 8).value = f"=SUM(H4:H{linha-1})"
    ws.cell(linha, 9).value = f"=IF(G{linha}=0,0,H{linha}/G{linha})"
    ws.cell(linha, 10).value = f"=IF(G{linha}=0,0,B{linha}/G{linha})"
    ws.cell(linha, 11).value = f"=IF(G{linha}=0,0,C{linha}/G{linha})"
    ws.cell(linha, 12).value = f"=IF(G{linha}=0,0,D{linha}/G{linha})"
    ws.cell(linha, 13).value = f"=IF(G{linha}=0,0,F{linha}/G{linha})"
    for col in range(1, 14):
        c = ws.cell(linha, col)
        c.fill = fill_resumo_total
        c.font = font_bold
        c.border = borda
        c.alignment = Alignment(horizontal="center")
        if col in [9, 10, 11, 12, 13]:
            c.number_format = "0.00%"
        if col == 8:
            c.number_format = "0.00"

    for col, largura in {"A": 24, "B": 18, "C": 14, "D": 24, "E": 14, "F": 14, "G": 14, "H": 20, "I": 14, "J": 14, "K": 14, "L": 24, "M": 14}.items():
        ws.column_dimensions[col].width = largura

    ws_nao = wb.create_sheet("NAO_ENCONTRADOS")
    ws_nao.append(["ABA DE ORIGEM", "CÓDIGO NÃO ENCONTRADO"])
    for item in nao_encontrados:
        ws_nao.append([item["aba"], item["codigo"]])
    for row in ws_nao.iter_rows():
        for c in row:
            c.border = borda
    ws_nao.column_dimensions["A"].width = 25
    ws_nao.column_dimensions["B"].width = 50

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)

    total_docs = sum(i["total"] for i in resumo_por_aba)
    total_contribuicao = sum(i["contribuicao"] for i in resumo_por_aba)
    indicadores = {
        "total_documentos": total_docs,
        "percentual_medido": (total_contribuicao / total_docs) if total_docs else 0,
        "contribuicao_total": total_contribuicao,
        "nao_encontrados": len(nao_encontrados),
        "por_aba": resumo_por_aba,
    }

    return saida.getvalue(), indicadores, nao_encontrados
